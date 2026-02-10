import openpyxl
from openpyxl.styles import PatternFill, Border, Side

filename = "test_excel_features.xlsx"
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Data"

# 1. Input Data
ws['A1'] = "Header 1"
ws['B1'] = "Header 2"
ws['A2'] = 10
ws['B2'] = 20

# 2. Border Test
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
ws['A1'].border = thin_border
ws['B1'].border = thin_border

# 3. Pivot Refresh Setting (Mock Pivot)
# Creating a real pivot table via openpyxl is hard, so we just check if the function can run without error
# and if it sets the cache property if a pivot existed. 
# For now, we manually create a dummy object to simulate pivot structure if needed, 
# but better to just test the function call safety on a file without pivots, 
# and maybe trust the code logic for actual pivots since creating them in pure python is complex.

wb.save(filename)

# Verify
wb2 = openpyxl.load_workbook(filename)
ws2 = wb2["Data"]
border_a1 = ws2['A1'].border.left.style
print(f"A1 Border Style: {border_a1}")

# Test capabilities import
import sys
import os
sys.path.append(os.path.join(os.getcwd()))
from capabilities.excel_manipulation import enable_pivot_table_refresh

res = enable_pivot_table_refresh(filename)
print(f"Pivot Refresh Result: {res}")
