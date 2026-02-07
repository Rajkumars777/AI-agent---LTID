import openpyxl
import polars as pl
from typing import List, Dict

def read_excel_data(file_path: str) -> List[Dict]:
    """Read Excel data using Polars for performance."""
    try:
        df = pl.read_excel(file_path)
        return df.to_dicts()
    except Exception as e:
        return [{"error": str(e)}]

def write_excel_data(file_path: str, data: List[Dict]):
    """Write data to Excel using Polars."""
    try:
        df = pl.DataFrame(data)
        df.write_excel(file_path)
        return f"Successfully wrote to {file_path}"
    except Exception as e:
        return f"Error writing to Excel: {str(e)}"

def update_cell_style(file_path: str, cell_id: str, color_hex: str):
    """Update cell style using OpenPyXL (Polars doesn't support styling well)."""
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        cell = ws[cell_id]
        from openpyxl.styles import PatternFill
        cell.fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
        wb.save(file_path)
        return f"Updated style for {cell_id}"
    except Exception as e:
        return f"Error updating style: {str(e)}"
