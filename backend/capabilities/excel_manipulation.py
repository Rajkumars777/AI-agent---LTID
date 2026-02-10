import polars as pl
import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter
import os

import os
import pandas as pd

def convert_xls_to_xlsx(filepath: str) -> str:
    """
    Converts a legacy .xls file to .xlsx using pandas + xlrd.
    Returns the new filepath if successful, or None if failed.
    """
    try:
        if not filepath.lower().endswith('.xls'):
            return filepath
        
        new_path = filepath + "x" # .xls -> .xlsx
        
        # Read with pandas (requires xlrd)
        df = pd.read_excel(filepath)
        
        # Write to xlsx (requires openpyxl)
        df.to_excel(new_path, index=False)
        return new_path
    except Exception as e:
        print(f"Conversion failed: {e}")
        return None

def read_sheet_data(filepath: str, sheet_name: str = None, fmt: str = "markdown") -> str:
    """
    Reads the first 10 rows of an Excel sheet using Polars for performance.
    Returns a markdown or html table string.
    """
    # print(f"DEBUG: Reading {filepath} with sheet_name={sheet_name}") 
    try:
        if not os.path.exists(filepath):
            return f"Error: File '{filepath}' not found."

        # Polars read_excel
        # If sheet_name is specific, use it. Otherwise read all and take the first.
        try:
            if sheet_name:
                df = pl.read_excel(filepath, sheet_name=sheet_name)
            else:
                # Read all sheets (returns dict {name: df}) or first sheet depending on version/engine
                data = pl.read_excel(filepath)
                if isinstance(data, dict):
                     if not data: return "Error: Excel file is empty."
                     # Take the first dataframe
                     df = list(data.values())[0]
                else:
                     df = data
        except Exception as read_err:
             # Fallback logic: Try explicit Sheet1 if automagic failed
             if "no matching sheets" in str(read_err).lower():
                  try:
                      df = pl.read_excel(filepath, sheet_name="Sheet1")
                  except:
                      return f"Error reading with Polars: {str(read_err)}"
             else:
                  return f"Error reading with Polars: {str(read_err)}"
        
        # Take head (limit to 1000 rows to prevent crashing the UI, but cover most "samples")
        preview = df.head(1000)
        
        if fmt == "html":
             # Use pandas to_html (polars doesn't have direct to_html? It does via to_pandas)
             return preview.to_pandas().to_html(index=False, classes="excel-table")
        else:
             return f"**Preview of '{os.path.basename(filepath)}'**:\n{preview.to_pandas().to_markdown(index=False)}"
    except Exception as e:
        return f"Error reading excel file: {str(e)}"

def write_cell(filepath: str, sheet_name: str, cell: str, value: str) -> str:
    """
    Writes a value to a specific cell (e.g., 'A1').
    """
    try:
        if not os.path.exists(filepath):
            return f"Error: File '{filepath}' not found."

        wb = openpyxl.load_workbook(filepath)
        if sheet_name:
            if sheet_name not in wb.sheetnames:
                 return f"Error: Sheet '{sheet_name}' not found. Available: {wb.sheetnames}"
            ws = wb[sheet_name]
        else:
            ws = wb.active

        ws[cell] = value
        
        # Helper: Try to convert to number if possible
        if str(value).replace('.', '', 1).isdigit():
             try:
                 if '.' in str(value): ws[cell] = float(value)
                 else: ws[cell] = int(value)
             except: pass
             
        wb.save(filepath)
        return f"Successfully wrote '{value}' to {cell} in {os.path.basename(filepath)}."
    except Exception as e:
        return f"Error writing to cell: {str(e)}"

def append_row(filepath: str, sheet_name: str, data: list | dict) -> str:
    """
    Appends a row of data. Supports list (positional) or dict (column header mapping).
    """
    try:
        if not os.path.exists(filepath):
            return f"Error: File '{filepath}' not found."

        wb = openpyxl.load_workbook(filepath)
        if sheet_name:
             if sheet_name not in wb.sheetnames:
                 return f"Error: Sheet '{sheet_name}' not found."
             ws = wb[sheet_name]
        else:
             ws = wb.active

        # Handle Dictionary (Column Mapping)
        if isinstance(data, dict):
            # 1. Read Headers from Row 1
            headers = {}
            max_col = ws.max_column
            for col_idx in range(1, max_col + 1):
                cell_val = ws.cell(row=1, column=col_idx).value
                if cell_val:
                    # Normalize: "First Name" -> "firstname"
                    headers[str(cell_val).strip().lower()] = col_idx
            
            # 2. Construct Ordered Row
            # If header not found in input, stick with None (empty cell)
            new_row = [None] * max_col
            
            # 3. Fuzzy match input keys to headers
            # Input: {"first name": "Raj", "age": 21}
            # Headers: {"firstname": 2, "age": 5}
            for key, val in data.items():
                clean_key = key.strip().lower().replace(" ", "").replace("_", "")
                
                # Check against headers (also normalized)
                match_col_idx = None
                for h_name, h_idx in headers.items():
                    clean_h = h_name.replace(" ", "").replace("_", "")
                    if clean_key == clean_h or clean_key in clean_h:
                        match_col_idx = h_idx
                        break
                
                if match_col_idx:
                    # Lists are 0-indexed, Cols are 1-indexed
                    if match_col_idx <= len(new_row):
                        new_row[match_col_idx - 1] = val
                    else:
                        # Extend if needed (unlikely if max_col is correct)
                        new_row.append(val)
            
            # Clean up trailing nones if necessary, but openpyxl handles it.
            ws.append(new_row)
            
        else:
            # Handle List (Positional)
            ws.append(data)

        wb.save(filepath)
        return f"Successfully appended row to {os.path.basename(filepath)}."
    except Exception as e:
        return f"Error appending row: {str(e)}"

def delete_row(filepath: str, sheet_name: str, row_index: int) -> str:
    """
    Deletes a specific row by index (1-based).
    """
    try:
        if not os.path.exists(filepath):
            return f"Error: File '{filepath}' not found."

        wb = openpyxl.load_workbook(filepath)
        if sheet_name:
             if sheet_name not in wb.sheetnames:
                 return f"Error: Sheet '{sheet_name}' not found."
             ws = wb[sheet_name]
        else:
             ws = wb.active

        ws.delete_rows(row_index)
        wb.save(filepath)
        return f"Successfully deleted row {row_index} in {os.path.basename(filepath)}."
    except Exception as e:
        return f"Error deleting row: {str(e)}"

def set_style(filepath: str, sheet_name: str, cell_range: str, bg_color: str = None, font_color: str = None, bold: bool = False, border: bool = False) -> str:
    """
    Applies styles to a cell or range.
    Colors should be Hex codes or mapped names.
    """
    try:
        if not os.path.exists(filepath):
            return f"Error: File '{filepath}' not found."

        wb = openpyxl.load_workbook(filepath)
        if sheet_name:
             ws = wb[sheet_name]
        else:
             ws = wb.active
        
        # Color Mapping
        color_map = {
            "red": "FF0000", "green": "00FF00", "blue": "0000FF", "yellow": "FFFF00",
            "black": "000000", "white": "FFFFFF", "gray": "CCCCCC", "orange": "FFA500"
        }
        
        if bg_color and bg_color.lower() in color_map: bg_color = color_map[bg_color.lower()]
        if font_color and font_color.lower() in color_map: font_color = color_map[font_color.lower()]
        
        fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid") if bg_color else None
        font = Font(color=font_color, bold=bold) if (font_color or bold) else None
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')) if border else None

        # Check if range or single cell
        if ":" in cell_range:
            rows = ws[cell_range]
            for row in rows:
                for cell in row:
                    if fill: cell.fill = fill
                    if font: cell.font = font
                    if thin_border: cell.border = thin_border
        else:
            cell = ws[cell_range]
            if fill: cell.fill = fill
            if font: cell.font = font
            if thin_border: cell.border = thin_border

        wb.save(filepath)
        return f"Applied styles to {cell_range} in {os.path.basename(filepath)}."
    except Exception as e:
        return f"Error styling cells: {str(e)}"

def enable_pivot_table_refresh(filepath: str) -> str:
    """
    Attempts to enable 'Refresh on Load' for all Pivot Tables in the workbook.
    """
    try:
        if not os.path.exists(filepath):
            return f"Error: File '{filepath}' not found."
            
        wb = openpyxl.load_workbook(filepath)
        
        count = 0
        for sheet in wb.worksheets:
            for pivot in sheet._pivots:
                pivot.cache.refreshOnLoad = True
                count += 1
                
        wb.save(filepath)
        if count > 0:
            return f"Enabled 'Refresh on Load' for {count} Pivot Tables."
        else:
            return "No Pivot Tables found to update."
            
    except Exception as e:
        return f"Error updating Pivot Table settings: {str(e)}"
