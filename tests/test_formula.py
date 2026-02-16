import openpyxl
from openpyxl.styles import PatternFill

filename = "test_formula.xlsx"
wb = openpyxl.Workbook()
ws = wb.active

# Input raw data
ws['A1'] = 10
ws['A2'] = 20

# Write formula
ws['A3'] = "=SUM(A1:A2)"

# Apply Style
fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
ws['A3'].fill = fill

wb.save(filename)

# Verify
wb2 = openpyxl.load_workbook(filename)
ws2 = wb2.active
val_a1 = ws2['A1'].value
val_a2 = ws2['A2'].value
formula_a3 = ws2['A3'].value
style_a3 = ws2['A3'].fill.start_color.index

print(f"A1: {val_a1}")
print(f"A2: {val_a2}")
print(f"A3 (Formula): {formula_a3}")
print(f"A3 Style Color: {style_a3}")
