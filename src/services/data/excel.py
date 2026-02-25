import polars as pl
import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter
import os
import pandas as pd
import re

def convert_xls_to_xlsx(filepath: str) -> str:
    """
    Converts a legacy .xls file to .xlsx using pandas + xlrd.
    Returns the new filepath if successful, or None if failed.
    """
    try:
        if not filepath.lower().endswith('.xls'):
            return filepath
        
        new_path = filepath + "x" # .xls -> .xlsx
        
        # Read with pandas (requires xlrd)
        df = pd.read_excel(filepath)
        
        # Write to xlsx (requires openpyxl)
        df.to_excel(new_path, index=False)
        return new_path
    except Exception as e:
        print(f"Conversion failed: {e}")
        return None

def read_sheet_data(filepath: str, sheet_name: str = None, fmt: str = "markdown") -> str:
    """
    Reads the first 10 rows of an Excel sheet using Polars for performance.
    Returns a markdown or html table string.
    """
    try:
        if not os.path.exists(filepath):
            return f"Error: File '{filepath}' not found."

        try:
            if sheet_name:
                df = pl.read_excel(filepath, sheet_name=sheet_name)
            else:
                data = pl.read_excel(filepath)
                if isinstance(data, dict):
                     if not data: return "Error: Excel file is empty."
                     df = list(data.values())[0]
                else:
                     df = data
        except Exception as read_err:
             if "no matching sheets" in str(read_err).lower():
                  try:
                      df = pl.read_excel(filepath, sheet_name="Sheet1")
                  except:
                      return f"Error reading with Polars: {str(read_err)}"
             else:
                  return f"Error reading with Polars: {str(read_err)}"
        
        preview = df.head(1000)
        
        if fmt == "html":
             return preview.to_pandas().to_html(index=False, classes="excel-table")
        else:
             return f"**Preview of '{os.path.basename(filepath)}'**:\n{preview.to_pandas().to_markdown(index=False)}"
    except Exception as e:
        return f"Error reading excel file: {str(e)}"

def write_cell(filepath: str, sheet_name: str, cell: str, value: str) -> str:
    """
    Writes a value to a specific cell (e.g., 'A1').
    """
    try:
        if not os.path.exists(filepath):
            return f"Error: File '{filepath}' not found."

        wb = openpyxl.load_workbook(filepath)
        if sheet_name:
            if sheet_name not in wb.sheetnames:
                 return f"Error: Sheet '{sheet_name}' not found. Available: {wb.sheetnames}"
            ws = wb[sheet_name]
        else:
            ws = wb.active

        ws[cell] = value
        
        if str(value).replace('.', '', 1).isdigit():
             try:
                 if '.' in str(value): ws[cell] = float(value)
                 else: ws[cell] = int(value)
             except: pass
             
        wb.save(filepath)
        return f"Successfully wrote '{value}' to {cell} in {os.path.basename(filepath)}."
    except Exception as e:
        return f"Error writing to cell: {str(e)}"

def append_row(filepath: str, sheet_name: str, data: list | dict) -> str:
    """
    Appends a row of data. Supports list (positional) or dict (column header mapping).
    """
    try:
        if not os.path.exists(filepath):
            return f"Error: File '{filepath}' not found."

        wb = openpyxl.load_workbook(filepath)
        ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

        if isinstance(data, dict):
            headers = {}
            max_col = ws.max_column
            for col_idx in range(1, max_col + 1):
                cell_val = ws.cell(row=1, column=col_idx).value
                if cell_val:
                    headers[str(cell_val).strip().lower()] = col_idx
            
            new_row = [None] * max_col
            for key, val in data.items():
                clean_key = key.strip().lower().replace(" ", "").replace("_", "")
                match_col_idx = None
                for h_name, h_idx in headers.items():
                    clean_h = h_name.replace(" ", "").replace("_", "")
                    if clean_key == clean_h or clean_key in clean_h:
                        match_col_idx = h_idx
                        break
                
                if match_col_idx:
                    if match_col_idx <= len(new_row):
                        new_row[match_col_idx - 1] = val
                    else:
                        new_row.append(val)
            ws.append(new_row)
        else:
            ws.append(data)

        wb.save(filepath)
        return f"Successfully appended row to {os.path.basename(filepath)}."
    except Exception as e:
        return f"Error appending row: {str(e)}"

def delete_row(filepath: str, sheet_name: str, row_index: int) -> str:
    """
    Deletes a specific row by index (1-based).
    """
    try:
        if not os.path.exists(filepath):
            return f"Error: File '{filepath}' not found."

        wb = openpyxl.load_workbook(filepath)
        ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

        ws.delete_rows(row_index)
        wb.save(filepath)
        return f"Successfully deleted row {row_index} in {os.path.basename(filepath)}."
    except Exception as e:
        return f"Error deleting row: {str(e)}"

def set_style(filepath: str, sheet_name: str, cell_range: str, bg_color: str = None, font_color: str = None, bold: bool = False, border: bool = False) -> str:
    """
    Applies styles to a cell or range. Supports 'headers', 'column names', 'row 1'.
    """
    try:
        if not os.path.exists(filepath):
            return f"Error: File '{filepath}' not found."

        wb = openpyxl.load_workbook(filepath)
        ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
        
        # 1. Resolve Special Targets
        target_range = str(cell_range).lower().strip()
        # Strip parentheses and junk
        target_range = re.sub(r'[\(\)]', '', target_range).strip()
        
        if any(h in target_range for h in ["header", "column", "name", "row 1"]):
            max_col = ws.max_column
            target_range = f"A1:{get_column_letter(max_col)}1"
            print(f"DEBUG: Resolved '{cell_range}' to header range {target_range}")
        
        # 2. Color Mapping
        target_range = target_range.replace("in ", "").strip()
        color_map = {
            "red": "FF0000", "green": "00FF00", "blue": "0000FF", "yellow": "FFFF00",
            "black": "000000", "white": "FFFFFF", "gray": "CCCCCC", "orange": "FFA500", "pink": "FFC0CB"
        }
        
        if bg_color and bg_color.lower() in color_map: bg_color = color_map[bg_color.lower()]
        if font_color and font_color.lower() in color_map: font_color = color_map[font_color.lower()]
        
        # Create Style Objects
        fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid") if bg_color else None
        font = Font(color=font_color, bold=bold) if (font_color or bold) else None
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')) if border else None

        # Apply to Range or Cell
        if ":" in target_range:
            rows = ws[target_range]
            if isinstance(rows, (list, tuple)):
                for row in rows:
                    if isinstance(row, (list, tuple)):
                        for cell in row:
                            if fill: cell.fill = fill
                            if font: cell.font = font
                            if thin_border: cell.border = thin_border
                    else:
                        cell = row
                        if fill: cell.fill = fill
                        if font: cell.font = font
                        if thin_border: cell.border = thin_border
            else:
                cell = rows
                if fill: cell.fill = fill
                if font: cell.font = font
                if thin_border: cell.border = thin_border
        else:
            cell = ws[target_range]
            if fill: cell.fill = fill
            if font: cell.font = font
            if thin_border: cell.border = thin_border

        wb.save(filepath)
        return f"Successfully applied styling to '{cell_range}' ({target_range}) in {os.path.basename(filepath)}."
    except Exception as e:
        return f"Error styling cells: {str(e)}"

def enable_pivot_table_refresh(filepath: str) -> str:
    """
    Attempts to enable 'Refresh on Load' for all Pivot Tables in the workbook.
    """
    try:
        if not os.path.exists(filepath):
            return f"Error: File '{filepath}' not found."
            
        wb = openpyxl.load_workbook(filepath)
        count = 0
        for sheet in wb.worksheets:
            for pivot in sheet._pivots:
                pivot.cache.refreshOnLoad = True
                count += 1
                
        wb.save(filepath)
        if count > 0:
            return f"Enabled 'Refresh on Load' for {count} Pivot Tables."
        else:
            return "No Pivot Tables found to update."
            
    except Exception as e:
        return f"Error updating Pivot Table settings: {str(e)}"
